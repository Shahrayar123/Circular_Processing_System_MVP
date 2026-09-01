"""The Excel working file.

Three sheets, following the shape of the workbook ABL supplied with their BRD:

    Summary        department-wise counts of New / Amendment / Deletion
    Proposed Tests the working rows
    Week MIS       every document received, actionable yes/no, action taken

The real file has six sheets and 78 columns. This is the subset that tells the story;
the columns are named the way ABL names them so the output is recognisable.

Formatting follows the BRD: additions in green text, deletions in red strike-through
text — text formatting, not cell fills.
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config, store

NAVY = "1F3864"
GREEN = "2E7A4F"
AMBER = "9C6F11"
RED = "B03A30"
GREY = "595959"

HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Sr #", 18), ("Audit Dep", 11), ("Test Type", 14), ("Amendment Type", 20),
    ("Strata", 20), ("Existing Test Code", 18), ("Existing Test", 52),
    ("New Test / Changes in Existing Test", 62), ("Exception Code", 16),
    ("New Exception / Changes in Existing Exception", 52), ("Risk Rating", 12),
    ("Root Cause", 24), ("Regulatory Violation", 18),
    ("Circular Name (Summary)", 40), ("Circular Reference Number", 26),
    ("Date of Circular", 16), ("Clause Reference", 16), ("Remarks", 46),
    # Makes it self-evident that this is the pre-review working file: every row carries
    # where it has reached in the review, so an unapproved file cannot be mistaken for
    # an approved one.
    ("Review Status", 20),
]


def _style_header(sheet, row: int, columns) -> None:
    for i, (name, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=i, value=name)
        cell.fill, cell.font, cell.border = HEAD_FILL, HEAD_FONT, BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(i)].width = width
    sheet.row_dimensions[row].height = 30


def _banner(sheet, columns: int, text: str) -> None:
    cell = sheet.cell(row=1, column=1, value=text)
    cell.font = Font(bold=True, size=10, color=AMBER)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)


# ====== SHEETS ======

def _summary(wb: Workbook, proposals: list[dict]) -> None:
    sheet = wb.active
    sheet.title = "Summary"
    sheet.sheet_view.showGridLines = False
    cell = sheet.cell(row=3, column=1, value="DEPARTMENT WISE SUMMARY OF AUDIT TESTS")
    cell.font = Font(bold=True, size=12, color=NAVY)

    headers = [("Audit Department (Full Name)", 42), ("Short Name", 14), ("New", 10),
               ("Amendment", 13), ("Deletion", 11), ("No action", 12), ("Total", 10)]
    _style_header(sheet, 5, headers)

    row = 6
    for code in config.DEPARTMENTS:
        rows = [p for p in proposals if p["department"] == code]
        counts = [sum(1 for p in rows if p["change_type"] == t)
                  for t in ("New", "Amendment", "Deletion", "No action")]
        values = [config.DEPARTMENT_NAMES[code], code, *counts, len(rows)]
        for i, value in enumerate(values, start=1):
            c = sheet.cell(row=row, column=i, value=value)
            c.border = BORDER
            c.font = Font(size=10)
        row += 1

    totals = ["TOTAL", "", *[sum(1 for p in proposals if p["change_type"] == t)
                             for t in ("New", "Amendment", "Deletion", "No action")],
              len(proposals)]
    for i, value in enumerate(totals, start=1):
        c = sheet.cell(row=row, column=i, value=value)
        c.font = Font(bold=True, size=10, color=NAVY)
        c.border = BORDER

    # ---- review position ----
    row += 3
    cell = sheet.cell(row=row, column=1, value="REVIEW STATUS OF THESE PROPOSALS")
    cell.font = Font(bold=True, size=12, color=NAVY)
    row += 2
    _style_header(sheet, row, [("Status", 42), ("Count", 14)])
    by_status = {}
    for p in proposals:
        key = p.get("status") or "Pending review"
        by_status[key] = by_status.get(key, 0) + 1
    row += 1
    for status, count in sorted(by_status.items()):
        for i, value in enumerate((status, count), start=1):
            c = sheet.cell(row=row, column=i, value=value)
            c.border = BORDER
            c.font = Font(size=10, color=(GREEN if status == "Approved" else GREY))
        row += 1
    note = sheet.cell(
        row=row + 1, column=1,
        value="This is the WORKING FILE for review. It intentionally contains proposals "
              "that have not yet been approved. Only fully approved changes appear in "
              "the eAudit export.")
    note.font = Font(italic=True, size=9, color=GREY)


def _proposed_tests(wb: Workbook, proposals: list[dict], docs: dict) -> None:
    sheet = wb.create_sheet("Proposed Tests")
    sheet.freeze_panes = "A4"
    _banner(sheet, len(COLUMNS), "Additions in green · Deletions in red strike-through")
    _style_header(sheet, 3, COLUMNS)

    row = 4
    for p in proposals:
        doc = docs.get(p["document_id"], {})
        change = p["change_type"]

        values = [
            p["sr_no"], p["department"], change, p["amendment_type"] or "",
            p["strata"] or "", p["target_test_code"] or "",
            p["existing_test_description"] or "",
            p["proposed_test_description"] or "",
            p["exception_code"] or "",
            p["proposed_exception_description"] or "",
            p["risk_rating"] or "", p["root_cause"] or "",
            "Yes" if (doc.get("source") or "").startswith("email") else "No",
            doc.get("title") or "", doc.get("filename") or "",
            doc.get("doc_date") or "", p["clause_ref"] or "", p["rationale"] or "",
            p.get("status") or "Pending review",
        ]

        for i, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=i, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # BRD §5.6 — text formatting, not fills
            if change == "Deletion":
                cell.font = Font(size=9, color=RED, strike=True)
            elif change == "New":
                cell.font = Font(size=9, color=GREEN)
            elif change == "Amendment":
                cell.font = Font(size=9, color=AMBER)
            else:
                cell.font = Font(size=9, color=GREY)
        row += 1


def _week_mis(wb: Workbook, docs: list[dict], proposals: list[dict]) -> None:
    sheet = wb.create_sheet("Week MIS")
    sheet.sheet_view.showGridLines = False
    _banner(sheet, 9, "MIS of documents received in the period under review")

    headers = [("Sr. #", 8), ("Circular Origination", 20), ("Circular Name", 46),
               ("Reference / File", 40), ("Date", 15), ("Received via", 20),
               ("Actionable (Yes/No)", 18), ("Action Taken", 24), ("Week", 12)]
    _style_header(sheet, 3, headers)

    year, week, _ = date.today().isocalendar()
    origin = {"folder": "ABL / SBP", "email-body": "RIA", "email-attachment": "RIA"}
    via = {"folder": "Circular directory", "email-body": "RIA email body",
           "email-attachment": "RIA email attachment"}

    row = 4
    for i, doc in enumerate(docs, start=1):
        made = [p for p in proposals if p["document_id"] == doc["id"]
                and p["change_type"] != "No action"]
        if doc["status"] == "duplicate":
            actionable, action = "—", "Duplicate — already processed"
        elif made:
            actionable, action = "Yes", f"Action performed — {len(made)} change(s) proposed"
        else:
            actionable, action = "No", "Action Not Required — for information only"

        values = [i, origin.get(doc["source"], "ABL"), doc["title"] or "",
                  doc["filename"], doc["doc_date"] or "", via.get(doc["source"], ""),
                  actionable, action, f"{week} ({year})"]
        for j, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=j, value=value)
            cell.border = BORDER
            cell.font = Font(size=9, color=GREY if doc["status"] == "duplicate" else "000000")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1


def build() -> Path:
    proposals = store.query(
        "SELECT p.*, c.clause_ref FROM proposals p "
        "JOIN clauses c ON c.id = p.clause_id ORDER BY p.id")
    documents = store.query("SELECT * FROM documents ORDER BY id")
    by_id = {d["id"]: d for d in documents}

    wb = Workbook()
    _summary(wb, proposals)
    _proposed_tests(wb, proposals, by_id)
    _week_mis(wb, documents, proposals)

    config.ensure_dirs()
    path = config.OUTPUT_DIR / "Audit_Checklist_Working_File.xlsx"
    wb.save(path)
    return path
