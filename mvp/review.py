"""The human review step, and the eAudit hand-off.

The agreed solution's process flow is:

    circular intake → AI analysis → Excel working file → HUMAN REVIEW → eAudit

Everything before "human review" was already in the MVP; this module adds the last two
stages so the demonstration matches the flow that was proposed.

Kept deliberately small. The delivered system has named users, roles, an append-only
audit log and version snapshots on every edit. Here there is one reviewer at level 1,
one approver at level 2, and a simple decision log — enough to show the shape of the
governance without building it.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import config, store

PENDING_L1 = "Pending review"
PENDING_L2 = "Pending approval"
APPROVED = "Approved"
REJECTED = "Rejected"
CHANGES = "Changes requested"

NAVY = "1F3864"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _log(conn, proposal_id: int, level: int, decision: str, note: str) -> None:
    store.insert(conn, "approvals", {
        "proposal_id": proposal_id, "level": level, "decision": decision,
        "note": note or "", "decided_at": datetime.now().isoformat(timespec="seconds"),
    })


# Which statuses a decision at each level may act on. A level-1 decision on something
# already past level 1 is not a valid action, and must not reach the decision log —
# a duplicate approval in an audit trail is a finding, not a cosmetic problem.
_ACCEPTS = {
    1: {PENDING_L1, CHANGES},
    2: {PENDING_L2},
}


def current_status(proposal_id: int) -> str:
    row = store.one("SELECT status FROM proposals WHERE id = ?", (proposal_id,))
    return (row or {}).get("status", PENDING_L1)


def decide(proposal_id: int, level: int, decision: str, note: str = "") -> str:
    """Record one review decision and return the proposal's new status.

    Level 1 approval advances to level 2; it does not finalise. Only level 2 approval
    marks a proposal Approved, and only Approved proposals can be exported.

    A decision that does not match the proposal's current state is refused: nothing is
    written and nothing is logged. That stops a double-click, a stale browser tab or a
    replayed request from putting two decisions on one proposal.
    """
    status_now = current_status(proposal_id)
    if status_now not in _ACCEPTS.get(level, set()):
        return status_now                       # not actionable at this level — no-op

    with store.connect() as conn:
        if decision == APPROVED:
            status = APPROVED if level >= 2 else PENDING_L2
            new_level = 2 if level == 1 else 2
        elif decision == REJECTED:
            status, new_level = REJECTED, level
        else:
            status, new_level = CHANGES, 1        # an edit sends it back to level 1

        field = "reviewer_note" if level == 1 else "approver_note"
        conn.execute(
            f"UPDATE proposals SET status = ?, current_level = ?, {field} = ? WHERE id = ?",
            (status, new_level, note, proposal_id))
        _log(conn, proposal_id, level, decision, note)
    return status


def decide_many(proposal_ids: list[int], level: int, decision: str, note: str = "") -> int:
    for proposal_id in proposal_ids:
        decide(proposal_id, level, decision, note)
    return len(proposal_ids)


def reset_all() -> None:
    # An eAudit file written before the reset would otherwise survive and still be
    # offered for download, implying approvals that no longer exist.
    stale = config.OUTPUT_DIR / "eAudit_BAC_Export.xlsx"
    if stale.exists():
        stale.unlink()
    with store.connect() as conn:
        conn.execute("UPDATE proposals SET status = ?, current_level = 1, "
                     "reviewer_note = NULL, approver_note = NULL", (PENDING_L1,))
        conn.execute("DELETE FROM approvals")


def history(proposal_id: int) -> list[dict]:
    return store.query(
        "SELECT level, decision, note, decided_at FROM approvals "
        "WHERE proposal_id = ? ORDER BY id", (proposal_id,))


# ====== eAUDIT HAND-OFF ======

EAUDIT_COLUMNS = [
    ("Test Code (for eAudit BAC)", 22), ("Test (for eAudit BAC)", 62),
    ("Exception Code (for eAudit BAC)", 24), ("Exception (for eAudit BAC)", 52),
    ("Risk Rating (for eAudit BAC)", 18), ("Circular Reference Number (for eAudit)", 34),
    ("Audit Department", 16), ("Change Type", 14),
    ("Approved by (level 2)", 22), ("Approved on", 20),
]


def export_eaudit() -> tuple[Path | None, list[str]]:
    """Write the eAudit hand-off file. APPROVED PROPOSALS ONLY.

    Returns (path, refusals). If anything in scope is unapproved the export still runs
    for the approved rows, but every unapproved item is named — the delivered system
    refuses outright, and the demo shows the same check being made.
    """
    approved = store.query(
        "SELECT p.*, d.filename, d.title FROM proposals p "
        "JOIN documents d ON d.id = p.document_id "
        "WHERE p.status = ? AND p.change_type != 'No action' ORDER BY p.id", (APPROVED,))

    unapproved = store.query(
        "SELECT sr_no, status FROM proposals WHERE status != ? AND change_type != 'No action'",
        (APPROVED,))
    refusals = [f"{r['sr_no']} — {r['status']}" for r in unapproved]

    if not approved:
        return None, refusals

    wb = Workbook()
    sheet = wb.active
    sheet.title = "eAudit BAC"
    for i, (name, width) in enumerate(EAUDIT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=i, value=name)
        cell.fill, cell.font, cell.border = HEAD_FILL, HEAD_FONT, BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[chr(64 + i)].width = width
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"

    for row_index, p in enumerate(approved, start=2):
        decision = store.one(
            "SELECT decided_at FROM approvals WHERE proposal_id = ? AND level = 2 "
            "ORDER BY id DESC LIMIT 1", (p["id"],))
        values = [
            p["target_test_code"] or p["sr_no"],
            p["proposed_test_description"] or p["existing_test_description"] or "",
            p["exception_code"] or "",
            p["proposed_exception_description"] or "",
            p["risk_rating"] or "", p["title"] or p["filename"] or "",
            p["department"] or "", p["change_type"],
            "approver (level 2)", (decision or {}).get("decided_at", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.border = BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    config.ensure_dirs()
    path = config.OUTPUT_DIR / "eAudit_BAC_Export.xlsx"
    wb.save(path)
    return path, refusals
